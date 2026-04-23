# ipi_v2.8 — FINAL PROFESSIONAL EDITION: Template + All file types work perfectly
import os
import sys
import threading
from datetime import datetime
import pandas as pd
import requests
from tkinter import filedialog, messagebox
import customtkinter as ctk
from PIL import Image, ImageTk
import ipaddress

# --- Folium ---
try:
    import folium
    from folium.plugins import HeatMap
    FOLIUM_AVAILABLE = True
except ImportError:
    FOLIUM_AVAILABLE = False

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

def is_valid_ip(ip_str):
    try:
        ipaddress.ip_address(str(ip_str).strip())
        return True
    except:
        return False

class IPProcessorApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("IP Intel Tool")
        self.geometry("900x600")
        self.resizable(False, False)

        self.input_file = None
        self.results_df = None
        self.output_dir = os.path.expanduser("~")

        self.setup_ui()
        self.after(500, self.load_logo)

    def setup_ui(self):
        ctk.CTkLabel(self, text="IP Intel Tool — Bulk IP address intelligence from ARIN", 
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=20)

        top_frame = ctk.CTkFrame(self)
        top_frame.pack(pady=10, padx=20, fill="x")

        ctk.CTkButton(top_frame, text="Instructions", command=self.show_instructions,
                      fg_color="#FDED03", hover_color="#FFB300", text_color="black",
                      font=ctk.CTkFont(weight="bold")).pack(side="left", padx=8)
        ctk.CTkButton(top_frame, text="Download Template", command=self.download_template,
                      fg_color="#9C27B0", hover_color="#7B1FA2").pack(side="left", padx=8)
        ctk.CTkButton(top_frame, text="Select File", command=self.select_file,
                      fg_color="#00A1F1", hover_color="#0078D4",
                      font=ctk.CTkFont(weight="bold")).pack(side="left", padx=8)

        self.file_label = ctk.CTkLabel(top_frame, text="No file selected", text_color="gray")
        self.file_label.pack(side="left", padx=20)

        ctk.CTkLabel(self, text="Supported file types: CSV • Microsoft Excel • Libre Office ODS", 
                     text_color="#00FFAA", font=ctk.CTkFont(size=11)).pack(pady=5)

        self.progress = ctk.CTkProgressBar(self, width=740)
        self.progress.pack(pady=20)
        self.progress.set(0)

        self.log = ctk.CTkTextbox(self, height=260, state="disabled", 
                                  font=ctk.CTkFont(family="Consolas", size=10))
        self.log.pack(padx=20, pady=(10, 60), fill="x")

        self.start_btn = ctk.CTkButton(self, text="Start Processing", state="disabled",
                                       fg_color="green", hover_color="darkgreen", height=40,
                                       font=ctk.CTkFont(size=14, weight="bold"), 
                                       command=self.start_processing)
        self.start_btn.pack(pady=15)

    def load_logo(self):
        try:
            base_path = sys._MEIPASS if getattr(sys, 'frozen', False) else os.path.dirname(__file__)
            logo_path = os.path.join(base_path, "IP_intel_logo.png")
            if os.path.exists(logo_path):
                img = Image.open(logo_path).resize((130, 130), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                label = ctk.CTkLabel(self, image=photo, text="")
                label.image = photo
                label.place(relx=1.0, rely=1.0, x=-30, y=-15, anchor="se")
        except: pass

    def log_msg(self, text):
        self.log.configure(state="normal")
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")
        self.update_idletasks()

    def show_instructions(self):
        messagebox.showinfo(
            "IP Intel Tool v2.8",
            "How to use this tool to find intel on your IP addresses\n\n"
            "1.) Always work from a copy of your data - NEVER the originals!\n\n"
            "2.) Click on the blue \"Select File\" button\n\n"
            "3.) Select your source file — can be .xls, .xlsx, .csv, or .ods\n\n"
            "4.) Click on the green \"Start Processing\" button\n\n"
            "5.) When processing completes, choose where to save your report\n\n"
            "6.) Open the report and map files to analyze your data"
        )

    def download_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"IP_Evidence_Template_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        if not path:
            return

        df = pd.DataFrame({
            "Timestamp": ["", "", "", "", "", "", "", "", "", ""],
            "IP": ["", "", "", "", "", "", "", "", "", ""]
        })

        try:
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="IP Evidence", startrow=3)

                worksheet = writer.sheets["IP Evidence"]
                from openpyxl.styles import Font, PatternFill, Alignment

                # Header styling
                header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
                header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
                for col in [1, 2]:
                    cell = worksheet.cell(row=4, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                worksheet.column_dimensions['A'].width = 24
                worksheet.column_dimensions['B'].width = 20

                # Title & Instructions
                title = worksheet.cell(1, 1, "IP EVIDENCE TEMPLATE")
                title.font = Font(size=18, bold=True, color="00FFAA")
                worksheet.merge_cells('A1:B1')
                title.alignment = Alignment(horizontal="center")

                inst = worksheet.cell(2, 1, "Enter one IP address per row with optional timestamp. IPv4 and IPv6 supported.")
                inst.font = Font(size=11, italic=True, color="AAAAAA")
                worksheet.merge_cells('A2:B2')

            messagebox.showinfo("Success", f"Professional template created!\n\n{os.path.basename(path)}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to create template:\n{e}")

    def select_file(self):
        self.input_file = filedialog.askopenfilename(
            title="Select evidence file",
            filetypes=[("All Supported", "*.csv *.xlsx *.xls *.ods"), ("All Files", "*.*")]
        )
        if self.input_file:
            self.file_label.configure(text=os.path.basename(self.input_file), text_color="white")
            self.start_btn.configure(state="normal")
            self.log_msg(f"Loaded: {os.path.basename(self.input_file)}")

    def load_input_file(self):
        ext = os.path.splitext(self.input_file)[1].lower()
        try:
            if ext == ".csv":
                return pd.read_csv(self.input_file)

            # Smart reading: detect if it's our template by checking first cell
            try:
                preview = pd.read_excel(self.input_file, nrows=1, engine="openpyxl" if ext != ".ods" else "odf")
                first_cell = str(preview.iloc[0, 0]).strip().upper()
                if "IP EVIDENCE TEMPLATE" in first_cell or first_cell == "IP EVIDENCE TEMPLATE":
                    # It's our template → skip 3 rows
                    return pd.read_excel(self.input_file, engine="openpyxl" if ext != ".ods" else "odf", skiprows=3)
            except:
                pass  # Fall back to normal read

            # Normal file → read from row 1
            engine = "odf" if ext == ".ods" else "openpyxl"
            return pd.read_excel(self.input_file, engine=engine)

        except Exception as e:
            raise ValueError(f"Could not read file:\n{e}")

    def process(self):
        self.start_btn.configure(state="disabled")
        self.progress.set(0)
        self.log_msg("Processing IPs...")
        try:
            df = self.load_input_file()

            ts_col = next((c for c in df.columns if any(x in str(c).lower() for x in ['time','date','stamp'])), None)
            ip_col = next((c for c in df.columns if 'ip' in str(c).lower()), None)

            if ip_col is None:
                raise ValueError("No IP column found. Please ensure a column contains 'IP' in the name.")

            df['IP_CLEAN'] = df[ip_col].astype(str).str.strip()
            df = df[df['IP_CLEAN'].apply(is_valid_ip)]
            df = df[df['IP_CLEAN'].notna() & (df['IP_CLEAN'] != 'nan') & (df['IP_CLEAN'] != '')]

            if df.empty:
                raise ValueError("No valid IP addresses found")

            if ts_col:
                df['Parsed_Time'] = df[ts_col].apply(self.parse_timestamp)

            df = df.drop_duplicates(subset='IP_CLEAN', keep='first')
            ips = df['IP_CLEAN'].tolist()
            total = len(ips)

            results = []
            for i, ip in enumerate(ips):
                self.log_msg(f"[{i+1}/{total}] {ip}")
                row = df[df['IP_CLEAN'] == ip].iloc[0]
                geo = self.get_geo(ip)
                arin = self.get_arin(ip)
                results.append({
                    'Timestamp': str(row.get(ts_col, '')) if ts_col else '',
                    'IP': ip, **geo, **arin
                })
                self.progress.set((i + 1) / total)

            final_df = pd.DataFrame(results)
            if 'Timestamp' in final_df.columns and final_df['Timestamp'].notna().any():
                final_df['sort'] = final_df['Timestamp'].apply(self.parse_timestamp)
                final_df = final_df.sort_values('sort', ascending=False, na_position='last').drop(columns='sort', errors='ignore')

            self.results_df = final_df
            self.log_msg(f"Complete! {total} IPs processed")
            self.after(300, self.save_results)

        except Exception as e:
            self.log_msg(f"ERROR: {e}")
            messagebox.showerror("Processing Failed", str(e))
        finally:
            self.start_btn.configure(state="normal")

    def save_results(self):
        if self.results_df is None or self.results_df.empty:
            self.log_msg("No data to save")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("LibreOffice", "*.ods"), ("CSV", "*.csv")],
            initialfile=f"IP_Intel_Report_{datetime.now():%Y%m%d_%H%M%S}"
        )
        if not save_path:
            self.log_msg("Save cancelled")
            return

        self.output_dir = os.path.dirname(save_path)

        try:
            if save_path.lower().endswith(".ods"):
                self.results_df.to_excel(save_path, engine="odf", index=False)
            elif save_path.lower().endswith(".xlsx"):
                self.results_df.to_excel(save_path, index=False)
            else:
                self.results_df.to_csv(save_path, index=False)
            self.log_msg(f"Report saved: {os.path.basename(save_path)}")

            map_path = os.path.join(self.output_dir, f"IP_Intel_Map_{datetime.now():%Y%m%d_%H%M%S}.html")
            if self.generate_map_report(self.results_df, map_path):
                self.log_msg(f"Map saved: {os.path.basename(map_path)}")
            else:
                self.log_msg("No location data — map not created")

            messagebox.showinfo("Success", 
                f"Processing Complete!\n\n"
                f"Report: {os.path.basename(save_path)}\n"
                f"Map: {os.path.basename(map_path) if os.path.exists(map_path) else 'Not generated'}\n\n"
                f"Files saved to:\n{self.output_dir}")

        except Exception as e:
            messagebox.showerror("Save Failed", f"Error:\n{e}")

    def generate_map_report(self, df, save_path):
        if not FOLIUM_AVAILABLE:
            return False
        valid = df.dropna(subset=['lat', 'lon'])
        if valid.empty:
            return False
        m = folium.Map(location=[valid['lat'].mean(), valid['lon'].mean()], zoom_start=4, tiles="CartoDB dark_matter")
        for _, r in valid.iterrows():
            popup = f"<b>{r['IP']}</b><br>{r['city']}, {r['country']}<br>{r['isp']}"
            folium.CircleMarker([r['lat'], r['lon']], radius=7, color="#00FF00", fill=True, fill_opacity=0.8,
                                popup=folium.Popup(popup, max_width=300)).add_to(m)
        HeatMap([[r['lat'], r['lon']] for _, r in valid.iterrows()], radius=15, blur=25).add_to(m)
        try:
            m.save(save_path)
            return True
        except:
            return False

    def parse_timestamp(self, ts):
        if pd.isna(ts) or not str(ts).strip():
            return None
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%m/%d/%Y %H:%M:%S", "%d/%m/%Y"]:
            try:
                return datetime.strptime(str(ts).strip(), fmt)
            except:
                continue
        return None

    def get_geo(self, ip):
        try:
            r = requests.get(f"http://ip-api.com/json/{ip}?fields=country,regionName,city,isp,org,lat,lon", timeout=12)
            j = r.json()
            if j.get("status") == "fail":
                raise Exception()
            return {
                'country': j.get('country','N/A'),
                'regionName': j.get('regionName','N/A'),
                'city': j.get('city','N/A'),
                'isp': j.get('isp','N/A'),
                'org': j.get('org','N/A'),
                'lat': j.get('lat'),
                'lon': j.get('lon')
            }
        except:
            return {'country':'N/A','regionName':'N/A','city':'N/A','isp':'N/A','org':'N/A','lat':None,'lon':None}

    def get_arin(self, ip):
        try:
            r = requests.get(f"https://whois.arin.net/rest/ip/{ip}", headers={'Accept': 'application/json'}, timeout=10)
            if r.status_code == 200:
                org = r.json()['net']['orgRef']['@name']
                return {'ARIN_Org': org, 'ARIN_NetRange': 'N/A'}
        except: pass
        return {'ARIN_Org': 'N/A', 'ARIN_NetRange': 'N/A'}

    def start_processing(self):
        threading.Thread(target=self.process, daemon=True).start()

if __name__ == "__main__":
    app = IPProcessorApp()
    app.mainloop()